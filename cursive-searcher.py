import os
import openpyxl
import shutil

def search_file(target_dir, search_text):
    """Recursive search for an HTML file containing the given search text in the target directory."""
    for root, _, files in os.walk(target_dir):
        for file_name in files:
            if file_name.lower().endswith(".html"):
                file_path = os.path.join(root, file_name)
                try:
                    with open(file_path, 'r', encoding='utf-8') as file:
                        if search_text in file.read():
                            print(f"Found '{search_text}' in: {file_path}")
                            return file_path
                except (UnicodeDecodeError, PermissionError):
                    pass
                # If the full text search is unsuccessful, check if the filename matches
                if search_text.lower() == os.path.splitext(file_name)[0].lower():
                    print(f"Found filename match '{file_name}' in: {file_path}")
                    return file_path
    return None

def update_excel_with_file_path(excel_file, target_dir, output_file):
    """Updates the copy of the Excel file with file paths for the cells in column A containing the search text."""
    # Check if the Excel file exists
    if not os.path.exists(excel_file):
        print(f"Error: Excel file '{excel_file}' not found.")
        return

    # Check if the target directory exists
    if not os.path.exists(target_dir):
        print(f"Error: Target directory '{target_dir}' not found.")
        return

    # Load the original workbook
    wb = openpyxl.load_workbook(excel_file)
    sheet = wb.active

    # Iterate through each row in column A (excluding the header row)
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1):
        cell_value = row[0].value
        if cell_value:
            # Search for the text in the target directory and get the file path if found
            print(f"Checking for '{cell_value}' in the target directory...")
            file_path = search_file(target_dir, cell_value)
            if file_path:
                # Update the corresponding cell in column B with the file path
                sheet.cell(row=row[0].row, column=2, value=file_path)
                print(f"Updated cell B{row[0].row} with file path: {file_path}")
            else:
                # If the text match search is unsuccessful, update the cell in column B with "File not found"
                sheet.cell(row=row[0].row, column=2, value="File not found")
                print(f"Updated cell B{row[0].row} with 'File not found'")

    # Save the modified Excel file
    wb.save(output_file)
    print("Search and update completed.")

if __name__ == "__main__":
    # Replace with the actual path to your Excel file, target directory, and output file
    excel_file_path = "CDN-Files.xlsx"
    target_directory = "downloaded_html"
    output_file_path = "CDN-Files-output.xlsx"

    # Check if both the Excel file and target directory exist before running the script
    if not os.path.exists(excel_file_path) or not os.path.exists(target_directory):
        print("Error: Excel file or target directory not found.")
    else:
        # Call the function to update the copy of the Excel file
        update_excel_with_file_path(excel_file_path, target_directory, output_file_path)